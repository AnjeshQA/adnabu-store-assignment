import openpyxl
import os


def get_data(sheetName):
    """
    Reads data from the Excel file and converts numeric values
    to clean strings (removing .0 from floats).
    """
    # 1. Construct the path to your excel file
    # This assumes your folder structure is ProjectRoot/Utilities/dataProvider.py
    # and ProjectRoot/excel/testdata.xlsx
    base_path = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base_path, "..", "excel", "testdata.xlsx")

    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        print(f"Error: The file {excel_path} was not found.")
        return []

    sheet = workbook[sheetName]
    main_list = []

    # Iterate through rows starting from the second row (skipping headers)
    for row in range(2, sheet.max_row + 1):
        data_list = []
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row, column=col).value

            if value is not None:
                # --- CLEANING LOGIC ---
                # Check if the value is a float (like 9000000000.0)
                if isinstance(value, float):
                    # Convert to int first to drop .0, then to string
                    value = str(int(value))
                else:
                    # Otherwise, just make it a string
                    value = str(value)

                data_list.append(value)

        # Only add to the main list if the row isn't empty
        if data_list:
            main_list.append(tuple(data_list))

    return main_list